from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
import json
from dhis.config import REPORTS_DIRECTORY

# Tracked entity instance report
# Order the report files names
def get_next_excel_filename(base: str = "tei_report") -> Path:
    i = 1
    while True:
        filename = REPORTS_DIRECTORY / f"{base}_{i}.xlsx"
        if not filename.exists():
            return filename
        i += 1

# Load information into the xlsx file
def write_tei(data):
    filename = get_next_excel_filename()
    names = data["column_names"]

    wb = Workbook()
    ws = wb.active
    ws.title = "tei report"

    headers = ["tei_id"] + names + ["Admission / Discharge"]
    ws.append(headers)

    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Define color styles
    styles = {
        "missing data": PatternFill(start_color="9e9e9e", end_color="9e9e9e", fill_type="solid"),
        "Same admission and discharge date": PatternFill(start_color="4caf50", end_color="4caf50", fill_type="solid"),
        "Admission date registered after discharge date": PatternFill(start_color="ff9800", end_color="ff9800", fill_type="solid"),
        "Discharge date different from admission date": PatternFill(start_color="f44336", end_color="f44336", fill_type="solid"),
    }
    white_font = Font(color="FFFFFF")

    # Add rows
    for r in data["rows"]:
        row = [r["tei_id"]] + r["values"] + [r["comparison"]]
        ws.append(row)
        # Apply style to last column
        comparison_text = r["comparison"]
        fill = styles.get(comparison_text)
        if fill:
            cell = ws.cell(row=ws.max_row, column=len(headers))  # last column
            cell.fill = fill
            cell.font = white_font
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(filename)
    print(f"Report saved as: {filename}")


# Even editing report
def write_event_xlsx(event_before, event_after):

    output_path = REPORTS_DIRECTORY / "event_edit_result.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Event Edit Result"

    ws.append(["Event ID", "Data Element", "Old Value", "New Value"])

    for before_dv in event_before.get("dataValues", []):
        de = before_dv["dataElement"]
        old_val = before_dv["value"]

        new_val = None
        for after_dv in event_after.get("dataValues", []):
            if after_dv["dataElement"] == de:
                new_val = after_dv["value"]
                break

        ws.append([event_before["event"], de, old_val, new_val])

    wb.save(output_path)
    print(f"XLSX saved to: {output_path}")


# Dashboard json clone
def write_dashboard_json(metadata, filename="cloned_dashboard.json"):
    output_path = REPORTS_DIRECTORY / filename

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)

    print(f"Dashboard clone preview saved to: {output_path}")


def write_dashboard_json(metadata, prefix="dashboard_clone"):
    i = 1
    while True:
        filename = f"{prefix}_{i}.json"
        output_path = REPORTS_DIRECTORY / filename
        if not output_path.exists():
            break
        i += 1

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)

    print(f"Dashboard clone preview saved to: {output_path}")
